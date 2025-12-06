# backend/reportes_router.py

from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from fastapi.responses import StreamingResponse
from datetime import datetime
from typing import Optional

import io
from openpyxl import Workbook

from modules.db import get_db
from modules.reportes_services import obtener_reporte_calidad

def normalizar_fecha_inicio(fecha: datetime) -> datetime:
    return datetime.combine(fecha.date(), datetime.min.time())

def normalizar_fecha_fin(fecha: datetime) -> datetime:
    return datetime.combine(fecha.date(), datetime.max.time())

router = APIRouter(
    prefix="/api/reportes",
    tags=["Reportes de Calidad"]
)

@router.get("/semanal")
def reporte_semanal(
    fecha_inicio: datetime = Query(..., description="Formato YYYY-MM-DD"),
    fecha_fin: datetime = Query(..., description="Formato YYYY-MM-DD"),
    linea_id: Optional[int] = None,
    producto_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """
    Devuelve el reporte semanal en JSON.
    """
    # 👇 aquí normalizas
    fecha_inicio = normalizar_fecha_inicio(fecha_inicio)
    fecha_fin = normalizar_fecha_fin(fecha_fin)

    datos = obtener_reporte_calidad(
        db=db,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        linea_id=linea_id,
        producto_id=producto_id
    )

    return datos

@router.get("/semanal/excel")
def exportar_reporte_semanal_excel(
    fecha_inicio: datetime = Query(..., description="Formato YYYY-MM-DD"),
    fecha_fin: datetime = Query(..., description="Formato YYYY-MM-DD"),
    linea_id: Optional[int] = None,
    producto_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """
    Exporta el reporte semanal en formato Excel (.xlsx).
    Usa los mismos datos que /api/reportes/semanal.
    """

    # 1) Obtener los mismos datos del reporte en JSON
    datos = obtener_reporte_calidad(
        db=db,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        linea_id=linea_id,
        producto_id=producto_id,
    )

    # 2) Crear libro de Excel
    wb = Workbook()

    # -------- Hoja 1: Resumen --------
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"

    ws_resumen.append(["Métrica", "Valor"])
    ws_resumen.append(["Total piezas", datos["total_piezas"]])
    ws_resumen.append(["Total defectos", datos["total_defectos"]])
    ws_resumen.append(["Porcentaje defectos (%)", datos["porcentaje_defectos"]])

    # -------- Hoja 2: Clasificación de defectos --------
    ws_clasif = wb.create_sheet(title="Clasificación")
    ws_clasif.append(["Categoría", "Cantidad"])

    for item in datos["clasificacion_defectos"]:
        ws_clasif.append([item["categoria"], item["cantidad"]])

    # -------- Hoja 3: Tendencia diaria --------
    ws_tend = wb.create_sheet(title="Tendencia")
    ws_tend.append(["Día", "Piezas", "Defectos", "Porcentaje defectos (%)"])

    for item in datos["tendencia"]:
        ws_tend.append([
            item["dia"],
            item["piezas"],
            item["defectos"],
            item["porcentaje_defectos"],
        ])

    # 3) Guardar en memoria
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_semanal_{fecha_inicio.date()}_{fecha_fin.date()}.xlsx"

    # 4) Devolver como archivo descargable
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        },
    )
