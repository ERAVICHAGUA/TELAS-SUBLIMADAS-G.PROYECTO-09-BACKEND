# backend/modules/export_service.py
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional


def generar_pdf_historial(
    registros: List[Dict],
    cliente: Optional[str] = None,
    fecha_inicio: Optional[datetime] = None,
    fecha_fin: Optional[datetime] = None,
    usuario: Optional[str] = None
) -> BytesIO:
    """
    Genera un PDF con el historial de calidad.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Estilo para título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    # Título
    story.append(Paragraph("Historial de Calidad por Cliente", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Metadatos
    meta_data = []
    if cliente:
        meta_data.append(["Cliente:", cliente])
    if fecha_inicio:
        meta_data.append(["Fecha Inicio:", fecha_inicio.strftime("%d/%m/%Y")])
    if fecha_fin:
        meta_data.append(["Fecha Fin:", fecha_fin.strftime("%d/%m/%Y")])
    if usuario:
        meta_data.append(["Usuario:", usuario])
    meta_data.append(["Fecha de Exportación:", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    meta_data.append(["Total de Registros:", str(len(registros))])
    
    meta_table = Table(meta_data, colWidths=[2*inch, 4*inch])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Tabla de datos
    if registros:
        # Encabezados
        data = [["ID", "Estado", "Fecha", "Responsable", "Lote", "Orden", "Cliente", "Tipo Producto", "Categoría", "Observaciones"]]
        
        # Datos
        for reg in registros:
            fecha_str = ""
            if reg.get("fecha_inspeccion"):
                try:
                    fecha_dt = datetime.fromisoformat(reg["fecha_inspeccion"].replace('Z', '+00:00'))
                    fecha_str = fecha_dt.strftime("%d/%m/%Y %H:%M")
                except:
                    fecha_str = str(reg.get("fecha_inspeccion", ""))
            
            data.append([
                str(reg.get("id", "")),
                reg.get("estado", ""),
                fecha_str,
                reg.get("responsable", "") or "",
                reg.get("lote", "") or "",
                reg.get("orden", "") or "",
                reg.get("cliente", "") or "",
                reg.get("tipo_producto", "") or "",
                reg.get("categoria", ""),
                (reg.get("observaciones", "") or "")[:50] + "..." if len(reg.get("observaciones", "") or "") > 50 else (reg.get("observaciones", "") or "")
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[0.4*inch, 0.7*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch, 1*inch, 1*inch, 1*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No se encontraron registros para los criterios seleccionados.", styles['Normal']))
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def generar_excel_historial(
    registros: List[Dict],
    cliente: Optional[str] = None,
    fecha_inicio: Optional[datetime] = None,
    fecha_fin: Optional[datetime] = None,
    usuario: Optional[str] = None
) -> BytesIO:
    """
    Genera un archivo Excel con el historial de calidad.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Calidad"
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    meta_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Título
    ws.merge_cells('A1:J1')
    ws['A1'] = "Historial de Calidad por Cliente"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align
    
    # Metadatos
    row = 3
    if cliente:
        ws[f'A{row}'] = "Cliente:"
        ws[f'B{row}'] = cliente
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    if fecha_inicio:
        ws[f'A{row}'] = "Fecha Inicio:"
        ws[f'B{row}'] = fecha_inicio.strftime("%d/%m/%Y")
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    if fecha_fin:
        ws[f'A{row}'] = "Fecha Fin:"
        ws[f'B{row}'] = fecha_fin.strftime("%d/%m/%Y")
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    if usuario:
        ws[f'A{row}'] = "Usuario:"
        ws[f'B{row}'] = usuario
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    ws[f'A{row}'] = "Fecha de Exportación:"
    ws[f'B{row}'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = "Total de Registros:"
    ws[f'B{row}'] = len(registros)
    ws[f'A{row}'].font = Font(bold=True)
    row += 2
    
    # Encabezados de tabla
    headers = ["ID", "Estado", "Fecha Inspección", "Responsable", "Lote", "Orden", "Cliente", "Tipo Producto", "Categoría", "Observaciones"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    row += 1
    
    # Datos
    for reg in registros:
        fecha_str = ""
        if reg.get("fecha_inspeccion"):
            try:
                fecha_dt = datetime.fromisoformat(reg["fecha_inspeccion"].replace('Z', '+00:00'))
                fecha_str = fecha_dt.strftime("%d/%m/%Y %H:%M")
            except:
                fecha_str = str(reg.get("fecha_inspeccion", ""))
        
        data_row = [
            reg.get("id", ""),
            reg.get("estado", ""),
            fecha_str,
            reg.get("responsable", "") or "",
            reg.get("lote", "") or "",
            reg.get("orden", "") or "",
            reg.get("cliente", "") or "",
            reg.get("tipo_producto", "") or "",
            reg.get("categoria", ""),
            reg.get("observaciones", "") or ""
        ]
        
        for col_num, value in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = value
            cell.alignment = left_align
            cell.border = border
        
        row += 1
    
    # Ajustar ancho de columnas
    column_widths = [8, 12, 18, 15, 12, 12, 20, 18, 15, 40]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

